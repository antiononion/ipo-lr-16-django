from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import EmailMessage
from django.core.paginator import Paginator
from django.db.models import Q
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import datetime

from rest_framework import viewsets, generics, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from .models import Product, Element, Cart, Category, Owner
from .serializers import (
    ProductSerializer, CategorySerializer, OwnerSerializer,
    CartSerializer, CartItemSerializer, UserSerializer,
    RegisterSerializer, ProfileSerializer,
)
from .permissions import IsAdminOrReadOnly


# ─────────────────────────────────────────────
#  ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ: получить корзину
# ─────────────────────────────────────────────

def get_customer_cart(request):
    cart, _ = Cart.objects.get_or_create(user=request.user)
    return cart


# ─────────────────────────────────────────────
#  ОСНОВНЫЕ СТРАНИЦЫ
# ─────────────────────────────────────────────

def hello_world(request):
    return render(request, 'index.html')


def author(request):
    return render(request, 'author.html')


def shop(request):
    products = Product.objects.all()
    categories = Category.objects.all()

    search = request.GET.get('search', '').strip()
    if search:
        products = products.filter(
            Q(title__icontains=search) | Q(description__icontains=search)
        )

    category_id = request.GET.get('category', '')
    if category_id:
        products = products.filter(category_id=category_id)

    paginator = Paginator(products, 9)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'shop.html', {
        'products': page_obj,
        'categories': categories,
        'page_obj': page_obj,
    })


def product_info(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'product_info.html', {'product': product})


# ─────────────────────────────────────────────
#  КОРЗИНА
# ─────────────────────────────────────────────

@login_required
def cart_detail(request):
    cart = get_customer_cart(request)
    elements = Element.objects.filter(cart=cart).select_related('product')
    total_price = sum(item.cost() for item in elements)
    return render(request, 'cart.html', {
        'cart': cart,
        'elements': elements,
        'total_price': total_price,
    })


@login_required
def cart_add(request, product_id):
    cart = get_customer_cart(request)
    product = get_object_or_404(Product, id=product_id)

    element, created = Element.objects.get_or_create(
        cart=cart,
        product=product,
        defaults={'number': 1}
    )
    if not created:
        # проверяем складской остаток перед увеличением
        if element.number < product.numberOF:
            element.number += 1
            element.save()
        else:
            messages.error(request, f'На складе только {product.numberOF} шт. товара «{product.title}».')
            return redirect('cart_detail')

    messages.success(request, f'«{product.title}» добавлен в корзину.')
    return redirect('cart_detail')


@login_required
def cart_remove(request, item_id):
    element = get_object_or_404(Element, id=item_id, cart__user=request.user)
    element.delete()
    messages.success(request, 'Товар удалён из корзины.')
    return redirect('cart_detail')


@login_required
def cart_update(request, item_id):
    element = get_object_or_404(Element, id=item_id, cart__user=request.user)
    if request.method == 'POST':
        try:
            new_number = int(request.POST.get('number', 1))
        except (ValueError, TypeError):
            new_number = 1

        if new_number > 0:
            element.number = new_number
            try:
                element.save()   # вызовет clean() → проверит остаток
            except Exception as e:
                messages.error(request, str(e))
        else:
            element.delete()
    return redirect('cart_detail')


# ─────────────────────────────────────────────
#  ГЕНЕРАЦИЯ EXCEL-ЧЕКА
# ─────────────────────────────────────────────

def _generate_receipt_excel(order_info: dict, elements) -> bytes:
    """
    Создаёт Excel-чек и возвращает байты файла.
    order_info: {'full_name', 'email', 'phone', 'address', 'comment', 'total_price', 'order_date'}
    elements: QuerySet[Element]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Чек заказа'

    # Цвета (лесная тема)
    pine_fill   = PatternFill('solid', fgColor='1E3A1D')
    moss_fill   = PatternFill('solid', fgColor='3B5E3A')
    leaf_fill   = PatternFill('solid', fgColor='D4EAD2')
    amber_fill  = PatternFill('solid', fgColor='C8813A')
    mist_fill   = PatternFill('solid', fgColor='F0F4EF')

    white_bold = Font(color='FFFFFF', bold=True, size=11)
    dark_bold  = Font(color='1E3A1D', bold=True, size=11)
    dark_reg   = Font(color='2C1A0E', size=11)
    amber_bold = Font(color='C8813A', bold=True, size=12)

    thin = Side(style='thin', color='8CB88A')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')

    # ── ЗАГОЛОВОК ──
    ws.merge_cells('A1:F1')
    ws['A1'] = '[#] CAMPING PROJECT — ЧЕК ЗАКАЗА'
    ws['A1'].font = Font(color='FFFFFF', bold=True, size=16)
    ws['A1'].fill = pine_fill
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Дата: {order_info["order_date"]}   |   Покупатель: {order_info["full_name"]}'
    ws['A2'].font = Font(color='D4EAD2', size=11)
    ws['A2'].fill = moss_fill
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 22

    # ── ДАННЫЕ ДОСТАВКИ ──
    ws.merge_cells('A3:F3')
    ws['A3'] = 'ДАННЫЕ ДОСТАВКИ'
    ws['A3'].font = white_bold
    ws['A3'].fill = moss_fill
    ws['A3'].alignment = center
    ws.row_dimensions[3].height = 20

    info_rows = [
        ('Email',     order_info['email']),
        ('Телефон',   order_info['phone'] or '—'),
        ('Адрес',     order_info['address']),
        ('Комментарий', order_info['comment'] or '—'),
    ]
    for i, (label, value) in enumerate(info_rows, start=4):
        ws.merge_cells(f'A{i}:B{i}')
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = dark_bold
        ws[f'A{i}'].fill = leaf_fill
        ws[f'A{i}'].alignment = left
        ws[f'A{i}'].border = border

        ws.merge_cells(f'C{i}:F{i}')
        ws[f'C{i}'] = value
        ws[f'C{i}'].font = dark_reg
        ws[f'C{i}'].alignment = left
        ws[f'C{i}'].border = border
        ws.row_dimensions[i].height = 20

    # ── ТОВАРЫ ──
    header_row = len(info_rows) + 4 + 1   # = 9
    ws.merge_cells(f'A{header_row}:F{header_row}')
    ws[f'A{header_row}'] = 'СОСТАВ ЗАКАЗА'
    ws[f'A{header_row}'].font = white_bold
    ws[f'A{header_row}'].fill = pine_fill
    ws[f'A{header_row}'].alignment = center
    ws.row_dimensions[header_row].height = 20

    col_headers = ['#', 'Товар', 'Категория', 'Кол-во', 'Цена за шт.', 'Сумма']
    col_letters = ['A', 'B', 'C', 'D', 'E', 'F']
    th_row = header_row + 1
    for col, hdr in zip(col_letters, col_headers):
        cell = ws[f'{col}{th_row}']
        cell.value = hdr
        cell.font = white_bold
        cell.fill = moss_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[th_row].height = 22

    data_start = th_row + 1
    total = 0
    for idx, item in enumerate(elements, start=1):
        r = data_start + idx - 1
        fill = mist_fill if idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        row_data = [
            idx,
            item.product.title,
            str(item.product.category),
            item.number,
            float(item.product.price),
            float(item.cost()),
        ]
        for col, val in zip(col_letters, row_data):
            cell = ws[f'{col}{r}']
            cell.value = val
            cell.font = dark_reg
            cell.fill = fill
            cell.border = border
            cell.alignment = right if col in ('D', 'E', 'F') else left
        total += float(item.cost())
        ws.row_dimensions[r].height = 20

    # ── ИТОГО ──
    total_row = data_start + len(list(elements))
    ws.merge_cells(f'A{total_row}:E{total_row}')
    ws[f'A{total_row}'] = 'ИТОГО К ОПЛАТЕ:'
    ws[f'A{total_row}'].font = Font(color='1E3A1D', bold=True, size=13)
    ws[f'A{total_row}'].fill = leaf_fill
    ws[f'A{total_row}'].alignment = right
    ws[f'A{total_row}'].border = border

    ws[f'F{total_row}'] = total
    ws[f'F{total_row}'].font = Font(color='C8813A', bold=True, size=14)
    ws[f'F{total_row}'].fill = leaf_fill
    ws[f'F{total_row}'].alignment = right
    ws[f'F{total_row}'].border = border
    ws[f'F{total_row}'].number_format = '#,##0.00 ₽'
    ws.row_dimensions[total_row].height = 28

    # Число формат цен
    for r in range(data_start, total_row):
        for col in ('E', 'F'):
            ws[f'{col}{r}'].number_format = '#,##0.00 ₽'

    # ── ШИРИНА КОЛОНОК ──
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
#  ОФОРМЛЕНИЕ ЗАКАЗА
# ─────────────────────────────────────────────

@login_required
def checkout(request):
    cart = get_customer_cart(request)
    elements = Element.objects.filter(cart=cart).select_related('product', 'product__category')
    total_price = sum(item.cost() for item in elements)

    # Если корзина пуста — перенаправить
    if not elements.exists():
        messages.info(request, 'Ваша корзина пуста. Добавьте товары перед оформлением заказа.')
        return redirect('catalog')

    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        email     = request.POST.get('email', '').strip()
        phone     = request.POST.get('phone', '').strip()
        address   = request.POST.get('address', '').strip()
        comment   = request.POST.get('comment', '').strip()

        # Простая валидация
        if not full_name or not email or not address:
            messages.error(request, 'Пожалуйста, заполните обязательные поля: Имя, Email, Адрес.')
            return render(request, 'checkout.html', {
                'elements': elements,
                'total_price': total_price,
            })

        order_date = datetime.datetime.now().strftime('%d.%m.%Y %H:%M')

        order_info = {
            'full_name':  full_name,
            'email':      email,
            'phone':      phone,
            'address':    address,
            'comment':    comment,
            'total_price': total_price,
            'order_date': order_date,
        }

        # Генерируем Excel-чек
        excel_bytes = _generate_receipt_excel(order_info, elements)

        # Отправляем email с вложением
        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            from email.mime.base import MIMEBase
            from email import encoders

            msg = MIMEMultipart()
            msg['Subject'] = '[Camping Project] Receipt'
            msg['From'] = settings.EMAIL_HOST_USER
            msg['To'] = email

            body_text = (
                f'Hello, {full_name}!\n\n'
                f'Order total: {total_price}\n'
                f'Delivery address: {address}\n\n'
                f'Receipt is attached.'
            )
            msg.attach(MIMEText(body_text, 'plain', 'utf-8'))

            filename = f'receipt_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(excel_bytes)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
            msg.attach(part)

            with smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT) as server:
                server.starttls()
                server.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)
                server.sendmail(settings.EMAIL_HOST_USER, email, msg.as_bytes())

        except Exception as e:
            messages.warning(
                request,
                f'Заказ оформлен, но отправить email не удалось: {e}'
            )

        # Уменьшаем склад (вычитаем купленное количество)
        for item in elements:
            product = item.product
            product.numberOF = max(0, product.numberOF - item.number)
            product.save()

        # Очищаем корзину
        elements.delete()

        # Переходим на страницу успеха
        return render(request, 'checkout_success.html', {
            'full_name':   full_name,
            'email':       email,
            'address':     address,
            'total_price': total_price,
        })

    # GET — показываем форму
    return render(request, 'checkout.html', {
        'elements':    elements,
        'total_price': total_price,
    })


# ─────────────────────────────────────────────
#  СПИСОК ТОВАРОВ (расширенный, с фильтрами)
# ─────────────────────────────────────────────

def product_list(request):
    products = Product.objects.all()

    query = request.GET.get('q', '').strip()
    if query:
        products = products.filter(
            Q(title__icontains=query) | Q(description__icontains=query)
        )

    category_id = request.GET.get('category', '')
    if category_id:
        products = products.filter(category_id=category_id)

    owner_id = request.GET.get('owner', '')
    if owner_id:
        products = products.filter(owner_id=owner_id)

    categories = Category.objects.all()
    owners = Owner.objects.all()

    return render(request, 'shop.html', {
        'products':   products,
        'categories': categories,
        'owners':     owners,
    })


# ─────────────────────────────────────────────
#  ЛИЧНЫЙ КАБИНЕТ
# ─────────────────────────────────────────────

@login_required
def profile_view(request):
    return render(request, 'profile.html')


# ─────────────────────────────────────────────
#  API views (DRF)
# ─────────────────────────────────────────────

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAdminOrReadOnly]


class OwnerViewSet(viewsets.ModelViewSet):
    queryset = Owner.objects.all()
    serializer_class = OwnerSerializer
    permission_classes = [IsAdminOrReadOnly]


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAdminOrReadOnly]


class CartViewSet(viewsets.ModelViewSet):
    queryset = Cart.objects.all()
    serializer_class = CartSerializer
    permission_classes = [permissions.IsAuthenticated]


class CartItemViewSet(viewsets.ModelViewSet):
    queryset = Element.objects.all()
    serializer_class = CartItemSerializer
    permission_classes = [permissions.IsAuthenticated]


class RegisterView(generics.CreateAPIView):
    serializer_class = RegisterSerializer
    permission_classes = [permissions.AllowAny]


class MeView(generics.RetrieveUpdateAPIView):
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        return self.request.user
